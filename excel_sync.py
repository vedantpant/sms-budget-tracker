"""
Excel Sync with Duplicate Avoidance
Syncs Supabase transactions to Excel while checking for duplicates
"""
import openpyxl
from config import Config
from supabase_client import SupabaseClient
from logger import log
from typing import List, Dict
from datetime import datetime


class ExcelSync:
    """Sync Supabase transactions to Excel with duplicate detection"""

    def __init__(self):
        self.supabase = SupabaseClient()
        self.excel_file = Config.EXCEL_FILE
        self.ws_name = 'Budget Tracking'

    def get_excel_transactions(self) -> List[Dict]:
        """Get all transactions from Excel"""
        transactions = []

        try:
            wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            ws = wb[self.ws_name]

            for row in range(12, ws.max_row + 1):
                date = ws.cell(row, 3).value
                tx_type = ws.cell(row, 4).value
                category = ws.cell(row, 5).value
                amount = ws.cell(row, 6).value
                details = ws.cell(row, 7).value

                if not date or not amount:
                    continue

                transactions.append({
                    'date': str(date),
                    'type': tx_type,
                    'category': category,
                    'amount': float(amount),
                    'details': details
                })

            wb.close()
            log.info(f"[Excel] Read {len(transactions)} transactions from Excel")
            return transactions

        except Exception as e:
            log.error(f"[Excel] Error reading Excel: {e}")
            return []

    def find_missing_in_excel(self) -> List[Dict]:
        """Find transactions in Supabase but not in Excel"""

        supabase_txns = self.supabase.get_all_transactions()
        excel_txns = self.get_excel_transactions()

        log.info(f"[Sync] Supabase: {len(supabase_txns)}, Excel: {len(excel_txns)}")

        # Create index of Excel transactions (amount + date)
        excel_index = set()
        for txn in excel_txns:
            # Use amount and date as unique key
            key = f"{txn['amount']}_{txn['date']}"
            excel_index.add(key)

        # Find missing ones
        missing = []
        for txn in supabase_txns:
            tx_date = txn.get('timestamp', '').split()[0] if txn.get('timestamp') else ''
            key = f"{txn['amount']}_{tx_date}"

            if key not in excel_index:
                missing.append(txn)

        log.info(f"[Sync] Found {len(missing)} transactions missing in Excel")
        return missing

    def add_to_excel(self, transactions: List[Dict]) -> int:
        """Add missing transactions to Excel"""

        added = 0

        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb[self.ws_name]

            # Find next empty row
            next_row = ws.max_row + 1

            for txn in transactions:
                try:
                    # Extract data
                    timestamp = txn.get('timestamp', '')
                    date_str = timestamp.split()[0] if timestamp else ''
                    amount = txn.get('amount', 0)
                    merchant = txn.get('merchant_name', 'Unknown')
                    tx_type = txn.get('type', 'Expenses')
                    category = txn.get('category', 'Uncategorized')

                    # Convert date format if needed
                    if date_str:
                        try:
                            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                            date_formatted = date_obj.strftime("%d-%m-%Y")
                        except:
                            date_formatted = date_str
                    else:
                        date_formatted = ''

                    # Add to Excel
                    ws.cell(next_row, 3).value = date_formatted       # Date
                    ws.cell(next_row, 4).value = tx_type             # Type
                    ws.cell(next_row, 5).value = category            # Category
                    ws.cell(next_row, 6).value = amount              # Amount
                    ws.cell(next_row, 7).value = merchant            # Details/Merchant

                    added += 1
                    next_row += 1

                    log.info(f"[Sync] Added to Excel: {merchant} - INR {amount}")

                except Exception as e:
                    log.error(f"[Sync] Error adding transaction to Excel: {e}")
                    continue

            wb.save(self.excel_file)
            wb.close()
            log.info(f"[Sync] Successfully added {added}/{len(transactions)} to Excel")

        except Exception as e:
            log.error(f"[Sync] Error saving Excel: {e}")

        return added

    def full_sync(self) -> Dict:
        """Complete sync: find missing and add to Excel"""

        log.info("[Sync] Starting full Excel sync...")
        print()
        print("=" * 90)
        print("EXCEL SYNC - FINDING & ADDING MISSING TRANSACTIONS")
        print("=" * 90)
        print()

        # Find missing
        missing = self.find_missing_in_excel()

        if not missing:
            print("No missing transactions found. Excel is up-to-date!")
            return {
                "status": "up_to_date",
                "missing_count": 0,
                "added_count": 0
            }

        print(f"Found {len(missing)} transactions missing in Excel")
        print(f"Adding to Excel...")
        print()

        # Add to Excel
        added = self.add_to_excel(missing)

        print()
        print("=" * 90)
        print("SYNC COMPLETE")
        print("=" * 90)
        print(f"Missing transactions: {len(missing)}")
        print(f"Successfully added:   {added}")
        print(f"Failed:               {len(missing) - added}")
        print()

        return {
            "status": "success",
            "missing_count": len(missing),
            "added_count": added,
            "failed_count": len(missing) - added
        }


if __name__ == "__main__":
    syncer = ExcelSync()
    result = syncer.full_sync()
