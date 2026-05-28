from config import Config
from logger import log
from utils import CircuitBreaker
from exceptions import SupabaseConnectionError
from supabase import create_client

class SupabaseClient:
    def __init__(self):
        self.url = Config.SUPABASE_URL
        self.key = Config.SUPABASE_KEY
        
        if not self.url or not self.key:
            raise SupabaseConnectionError("Missing SUPABASE credentials in .env")
        
        self.client = create_client(self.url, self.key)

        self.circuit_breaker = CircuitBreaker(
            name='supabase',
            threshold=5,
            timeout=300
            )
        
        log.info("Supabase client initialized")

    
    def insert_sms_raw(self, sms_body: str, sender: str):
        from utils import generate_sms_hash
        from exceptions import DuplicateSMSError, SupabaseConnectionError

        sms_hash = generate_sms_hash(sms_body)

        try:
            response = self.client.table('sms_raw').insert({
                'sms_body': sms_body,
                'sender': sender,
                'sms_hash': sms_hash,
                'status': 'new',
                'source': 'manual'      
                }).execute()

            log.info(f"SMS inserted: {sms_hash[:8]}...")
            return response.data[0]
        
        except Exception as e:
            if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
                raise DuplicateSMSError(f"SMS already exists: {sms_hash[:8]}")
            raise SupabaseConnectionError(f"Insert failed: {e}")
        
    
    def get_category_map(self):
        try:
            response = self.client.table('category_map').select('*').execute()

            result = {}

            for row in response.data:
                result[row['merchant_name']] = {
                    'category': row['category'],
                    'type': row['type']
                    }

            log.info(f"Category map loaded: {len(result)} merchants")
            return result
        
        except Exception as e:
            raise SupabaseConnectionError(f"get_category_map failed: {e}")
        
    def insert_transaction(self, data: dict):
        from exceptions import DuplicateTransactionError, SupabaseConnectionError

        try:
            reponse = self.client.table('transactions').insert(
                {
                    'sms_id': data.get('sms_id'),
                    'amount': data.get('amount'),
                    'merchant_name': data['merchant_name'],
                    'type': data['type'],
                    'category': data['category'],
                    'transaction_id': data['transaction_id'],
                    'timestamp': data['timestamp'],
                    'account_number': data.get('account_number', '')
                }
            ).execute()

            log.info(f"Transaction inserted: {data['merchant_name']} ₹{data['amount']}")
            return reponse.data[0]
        
        except Exception as e:
            if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
                raise DuplicateTransactionError(f"Txn exists: {data['transaction_id']}")
            raise SupabaseConnectionError(f"insert_transaction failed: {e}")
        
    def migrate_category_map(self) -> dict:
        import json
        from config import Config

        with open(Config.CATEGORY_MAP_FILE, 'r') as f:
            data = json.load(f)
        
        inserted, skipped = 0, 0

        for merchant, values in data.items():
            try:
                self.client.table('category_map').insert(
                    {
                        'merchant_name': merchant,
                        'category': values['category'],
                        'type': values['type'],
                        'source': 'import'
                    }
                ).execute()
                inserted += 1
            except Exception as e:
                if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
                    skipped += 1
                else:
                    log.error(f"Failed: {merchant} - {e}")

        log.info(f"Migration done: {inserted} inserted, {skipped} skipped")
        return {'inserted': inserted, 'skipped': skipped}
    

    def migrate_excel(self) -> dict:
        import openpyxl
        from config import Config

        wb = openpyxl.load_workbook(Config.EXCEL_FILE, data_only=True)
        ws = wb['Budget Tracking']

        inserted, skipped, failed = 0, 0, 0

        for row in range(12, ws.max_row + 1):
            date    = ws.cell(row, 3).value
            txn_type = ws.cell(row, 4).value
            category = ws.cell(row, 5).value
            amount   = ws.cell(row, 6).value
            details  = ws.cell(row, 7).value

            if not date or not amount:
                continue

        # Parse "Merchant | TxnID" from details column
            if details and '|' in str(details):
                parts = str(details).split('|')
                merchant = parts[0].strip()
                txn_id   = parts[1].strip()
            else:
                merchant = str(details).strip() if details else 'Unknown'
                txn_id   = f"EXCEL_{row}"

            try:
                self.client.table('transactions').insert({
                    'amount':        float(amount),
                    'merchant_name': merchant,
                    'type':          txn_type,
                    'category':      category,
                    'transaction_id': txn_id,
                    'timestamp':     date.isoformat(),
                }).execute()
                inserted += 1
            except Exception as e:
                if 'duplicate' in str(e).lower() or 'unique' in str(e).lower():
                    skipped += 1
                else:
                    log.error(f"Row {row} failed: {e}")
                    failed += 1

        log.info(f"Excel migration: {inserted} inserted, {skipped} skipped, {failed} failed")
        return {'inserted': inserted, 'skipped': skipped, 'failed': failed}
            


if __name__ == "__main__":
    try:
        client = SupabaseClient()
        print(f"URL: {client.url}")
        print(f"Circuit Breaker: {client.circuit_breaker.name}")
    except Exception as e:
        print(f"Error: {e}")
        exit()

        # result = client.insert_sms_raw(
        #     sms_body="INR 500.00 debited from A/c XX1234 on 28-05-26 UPI/123456/Zomato",
        #     sender="AXISBK"
        #     )
        
        # print(f"✓ Inserted ID: {result['id']}")
    try:
        cat_map = client.get_category_map()
        print(f"Total merchants: {len(cat_map)}")
        print(f"Sample: {list(cat_map.items())[:2]}")
    except Exception as e:
        print(f"Error: {e}")

    try:
        txn = client.insert_transaction({
        'sms_id': None,
        'amount': 500.00,
        'merchant_name': 'Zomato',
        'type': 'Expenses',
        'category': 'Food Outside',
        'transaction_id': 'UPI123456789',
        'timestamp': '2026-05-28T20:44:00',
        'account_number': 'XX1234'
        })
        print(f"✓ Txn inserted: {txn['id']}")

    except Exception as e:
        print(f"Error: {e}")

    # result = client.migrate_category_map()
    # print("Migration result:", result)

    result = client.migrate_excel()
    print("Excel migration result:", result)
