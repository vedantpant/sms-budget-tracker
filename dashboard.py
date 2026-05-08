import streamlit as st
import openpyxl
from collections import defaultdict
import plotly.express as px

st.divider()

st.set_page_config(page_title="Budget Tracker", page_icon="💰", layout="wide")
st.title("💰 SMS Budget Tracker Dashboard")

# Load data (same logic as summary_report.py)
@st.cache_data
def load_transactions():
    wb = openpyxl.load_workbook("Ultimate Personal Budget Manager.xlsx", data_only=True)
    ws = wb["Budget Tracking"]
    transactions = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=3, max_col=7, values_only=True):
        date, type_, category, amount, merchant = row
        if date and amount and isinstance(amount, (int, float)):
            transactions.append({
                "date": date,
                "type": type_,
                "category": category,
                "amount": float(amount),
                "merchant": merchant.split("|")[0].strip() if merchant else "Unknown"
            })
    return transactions

transactions = load_transactions()
st.sidebar.success(f"Loaded {len(transactions)} transactions")

# Month selector
months = sorted(set(t["date"].strftime("%Y-%m") for t in transactions))
selected_month = st.sidebar.selectbox("Select month", months, index=len(months)-2)

# Filter transactions for selected month
filtered = [t for t in transactions if t["date"].strftime("%Y-%m") == selected_month]
st.sidebar.info(f"{len(filtered)} transactions in {selected_month}")

# Calculate stats for selected month
expenses = [t for t in filtered if t["type"] == "Expenses"]
income = [t for t in filtered if t["type"] == "Income"]
savings = [t for t in filtered if t["type"] == "Savings"]

total_exp = sum(t["amount"] for t in expenses)
total_inc = sum(t["amount"] for t in income)
total_sav = sum(t["amount"] for t in savings)

# Metric cards
col1, col2, col3, col4 = st.columns(4)
col1.metric("Total Expenses", f"₹{total_exp:,.0f}")
col2.metric("Total Income", f"₹{total_inc:,.0f}")
col3.metric("Total Savings", f"₹{total_sav:,.0f}")
col4.metric("Transactions", len(filtered))

st.divider()

# Category breakdown + Top merchants side by side
left, right = st.columns(2)

with left:
    st.subheader("🏷️ Category breakdown")
    cat_data = defaultdict(float)
    for t in expenses:
        cat_data[t["category"]] += t["amount"]
    
    sorted_cats = sorted(cat_data.items(), key=lambda x: -x[1])
    for cat, amt in sorted_cats:
        pct = (amt / total_exp * 100) if total_exp else 0
        st.text(f"{cat:25s}  ₹{amt:>10,.0f}  ({pct:.1f}%)")
        st.progress(pct / 100)

with right:
    st.subheader("🏪 Top 5 merchants")
    merch_data = defaultdict(lambda: {"amount": 0, "count": 0})
    for t in expenses:
        merch_data[t["merchant"]]["amount"] += t["amount"]
        merch_data[t["merchant"]]["count"] += 1
    
    top5 = sorted(merch_data.items(), key=lambda x: -x[1]["amount"])[:5]
    for i, (name, data) in enumerate(top5, 1):
        avg = data["amount"] / data["count"]
        st.text(f"{i}. {name[:25]:25s}")
        st.text(f"   ₹{data['amount']:,.0f} ({data['count']} txns, avg ₹{avg:,.0f})")
        st.text("")

# Plotly pie chart + bar chart
chart_left, chart_right = st.columns(2)

with chart_left:
    st.subheader("📊 Expense distribution")
    if cat_data:
        fig = px.pie(
            names=list(cat_data.keys()),
            values=list(cat_data.values()),
            hole=0.4
        )
        fig.update_layout(height=350, margin=dict(t=20, b=20, l=20, r=20))
        st.plotly_chart(fig, use_container_width=True)

with chart_right:
    st.subheader("📈 Monthly expense trend")
    monthly_exp = {}
    for t in transactions:
        if t["type"] == "Expenses":
            m = t["date"].strftime("%Y-%m")
            monthly_exp[m] = monthly_exp.get(m, 0) + t["amount"]
    
    fig2 = px.bar(
        x=list(sorted(monthly_exp.keys())),
        y=[monthly_exp[m] for m in sorted(monthly_exp.keys())],
        labels={"x": "Month", "y": "Expenses (₹)"}
    )
    fig2.update_layout(height=350, margin=dict(t=20, b=20, l=20, r=20))
    st.plotly_chart(fig2, use_container_width=True)

# AI Insights
st.divider()
st.subheader("🤖 AI Insights")

if st.button("Generate AI Insights", type="primary"):
    import requests
    with st.spinner("Ollama se insights generate ho rahe hain..."):
        report_data = f"""
Month: {selected_month}
Total Expenses: ₹{total_exp:,.0f}
Total Income: ₹{total_inc:,.0f}
Total Savings: ₹{total_sav:,.0f}
Top categories: {dict(sorted_cats[:5])}
Top merchants: {dict((n, d['amount']) for n, d in top5)}
"""
        try:
            response = requests.post("http://localhost:11434/api/generate", json={
                "model": "llama3.1:8b",
                "prompt": f"""You are a personal finance advisor for an Indian professional in Bengaluru.
Analyze this monthly spending and give 3 specific, actionable insights with ₹ amounts.
Be direct, mention specific merchants/categories.

{report_data}

Give exactly 3 bullet points, each 1-2 lines. No intro, no outro.""",
                "stream": False
            }, timeout=30)
            insights = response.json()["response"]
            st.info(insights)
        except:
            st.error("Ollama not running! Start with: ollama serve")

st.divider()
st.subheader("📥 Export report")

if st.button("Generate HTML Report"):
    # Build category rows
    cat_rows = ""
    for cat, amt in sorted_cats:
        pct = (amt / total_exp * 100) if total_exp else 0
        width = max(pct, 2)
        cat_rows += f"""
        <tr>
            <td>{cat}</td>
            <td style="width:50%"><div style="background:#3b82f6;height:20px;border-radius:4px;width:{width}%"></div></td>
            <td style="text-align:right;font-weight:600">₹{amt:,.0f}</td>
            <td style="text-align:right;color:#888">{pct:.1f}%</td>
        </tr>"""

    # Build merchant rows
    merch_rows = ""
    for i, (name, data) in enumerate(top5, 1):
        avg = data["amount"] / data["count"]
        merch_rows += f"""
        <tr>
            <td>{i}</td>
            <td>{name}</td>
            <td style="text-align:right;font-weight:600">₹{data['amount']:,.0f}</td>
            <td style="text-align:right;color:#888">{data['count']} txns, avg ₹{avg:,.0f}</td>
        </tr>"""

    # Monthly data for chart
    monthly_exp = {}
    for t in transactions:
        if t["type"] == "Expenses":
            m = t["date"].strftime("%Y-%m")
            monthly_exp[m] = monthly_exp.get(m, 0) + t["amount"]
    chart_labels = sorted(monthly_exp.keys())
    chart_values = [round(monthly_exp[m]) for m in chart_labels]

    html_content = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Budget Report - {selected_month}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,sans-serif; background:#f8f9fa; padding:2rem; color:#222; }}
  .container {{ max-width:900px; margin:0 auto; }}
  h1 {{ font-size:28px; margin-bottom:8px; }}
  .subtitle {{ color:#666; margin-bottom:2rem; }}
  .metrics {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-bottom:2rem; }}
  .metric {{ background:#fff; border-radius:12px; padding:1.2rem; border:1px solid #eee; }}
  .metric-label {{ font-size:13px; color:#888; }}
  .metric-value {{ font-size:26px; font-weight:700; margin-top:4px; }}
  .section {{ background:#fff; border-radius:12px; padding:1.5rem; border:1px solid #eee; margin-bottom:1.5rem; }}
  .section h2 {{ font-size:18px; margin-bottom:1rem; }}
  table {{ width:100%; border-collapse:collapse; }}
  td {{ padding:8px 4px; font-size:14px; }}
  .charts {{ display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:1.5rem; }}
  .insight {{ background:#eff6ff; border-radius:8px; padding:1rem; margin-bottom:8px; font-size:14px; color:#1e40af; }}
  .footer {{ text-align:center; color:#aaa; font-size:12px; margin-top:2rem; }}
</style>
</head>
<body>
<div class="container">
  <h1>💰 Budget Report — {selected_month}</h1>
  <p class="subtitle">Generated by SMS Budget Tracker (AI-powered)</p>

  <div class="metrics">
    <div class="metric"><div class="metric-label">Total expenses</div><div class="metric-value" style="color:#e53e3e">₹{total_exp:,.0f}</div></div>
    <div class="metric"><div class="metric-label">Total income</div><div class="metric-value" style="color:#38a169">₹{total_inc:,.0f}</div></div>
    <div class="metric"><div class="metric-label">Total savings</div><div class="metric-value" style="color:#3182ce">₹{total_sav:,.0f}</div></div>
    <div class="metric"><div class="metric-label">Transactions</div><div class="metric-value">{len(filtered)}</div></div>
  </div>

  <div class="charts">
    <div class="section"><h2>📊 Expense distribution</h2><canvas id="pieChart"></canvas></div>
    <div class="section"><h2>📈 Monthly trend</h2><canvas id="barChart"></canvas></div>
  </div>

  <div class="section"><h2>🏷️ Category breakdown</h2><table>{cat_rows}</table></div>
  <div class="section"><h2>🏪 Top 5 merchants</h2><table>{merch_rows}</table></div>

  <p class="footer">SMS Budget Tracker • AI-powered by Ollama • Generated {selected_month}</p>
</div>
<script>
new Chart(document.getElementById('pieChart'), {{
  type: 'doughnut',
  data: {{ labels: {list(cat_data.keys())}, datasets: [{{ data: {list(cat_data.values())}, backgroundColor: ['#e53e3e','#3182ce','#38a169','#d69e2e','#805ad5','#dd6b20','#319795','#e53e3e','#718096'] }}] }},
  options: {{ responsive: true, plugins: {{ legend: {{ position: 'bottom', labels: {{ font: {{ size: 11 }} }} }} }} }}
}});
new Chart(document.getElementById('barChart'), {{
  type: 'bar',
  data: {{ labels: {chart_labels}, datasets: [{{ label: 'Expenses', data: {chart_values}, backgroundColor: '#3b82f6' }}] }},
  options: {{ responsive: true, plugins: {{ legend: {{ display: false }} }}, scales: {{ y: {{ beginAtZero: true }} }} }}
}});
</script>
</body>
</html>"""

    # Save and offer download
    filename = f"budget_report_{selected_month}.html"
    st.download_button(
        label=f"⬇️ Download {filename}",
        data=html_content,
        file_name=filename,
        mime="text/html"
    )
    st.success(f"Report ready! Click above to download.")