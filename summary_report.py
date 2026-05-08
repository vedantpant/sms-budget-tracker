import openpyxl
from collections import defaultdict
import requests
import json

wb = openpyxl.load_workbook("Ultimate Personal Budget Manager.xlsx", data_only=True)
ws = wb["Budget Tracking"]

transactions = []
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=3, max_col=7, values_only=True):
    date, type_, category, amount, merchant = row
    if date and amount and isinstance(amount, (int, float)):
        transactions.append({
            "date": date,
            "type": type_,
            "category": category,
            "amount": float(amount),
            "merchant": merchant
        })

print(f"Total transactions read: {len(transactions)}")
print(f"First 3: {transactions[:3]}")

monthly = defaultdict(list)
for t in transactions:
    key = t["date"].strftime("%Y-%m")
    monthly[key].append(t)

print(f"\nMonths found: {len(monthly)}")
for month in sorted(monthly.keys()):
    txns = monthly[month]
    total_expense = sum(t["amount"] for t in txns if t["type"] == "Expenses")
    total_income = sum(t["amount"] for t in txns if t["type"] == "Income")
    total_savings = sum(t["amount"] for t in txns if t["type"] == "Savings")
    print(f"{month}: {len(txns)} txns | Expenses: ₹{total_expense:,.0f} | Income: ₹{total_income:,.0f} | Savings: ₹{total_savings:,.0f}")


# Latest complete month (skip current incomplete month)
latest_month = sorted(monthly.keys())[-2]  # -2 = last complete month
txns = monthly[latest_month]

print(f"\n{'='*50}")
print(f"📊 Detailed Report: {latest_month}")
print(f"{'='*50}")

# Category breakdown (Expenses only)
cat_totals = defaultdict(float)
cat_counts = defaultdict(int)
for t in txns:
    if t["type"] == "Expenses":
        cat_totals[t["category"]] += t["amount"]
        cat_counts[t["category"]] += 1

print(f"\n🏷️ Category Breakdown:")
total_expense = sum(cat_totals.values())
for cat, amt in sorted(cat_totals.items(), key=lambda x: -x[1]):
    pct = (amt / total_expense * 100) if total_expense else 0
    bar = "█" * int(pct / 3)
    print(f"  {cat:25s} ₹{amt:>10,.0f}  ({pct:4.1f}%)  {bar}  [{cat_counts[cat]} txns]")

# Top 5 merchants
merchant_totals = defaultdict(float)
merchant_counts = defaultdict(int)
for t in txns:
    if t["type"] == "Expenses":
        name = t["merchant"].split("|")[0].strip()
        merchant_totals[name] += t["amount"]
        merchant_counts[name] += 1

print(f"\n🏪 Top 5 Merchants:")
for i, (merch, amt) in enumerate(sorted(merchant_totals.items(), key=lambda x: -x[1])[:5], 1):
    avg = amt / merchant_counts[merch]
    print(f"  {i}. {merch:25s} ₹{amt:>10,.0f}  ({merchant_counts[merch]} txns, avg ₹{avg:,.0f})")

# Month-over-month comparison
months_sorted = sorted(monthly.keys())
if len(months_sorted) >= 3:
    curr = months_sorted[-2]  # latest complete
    prev = months_sorted[-3]  # previous month
    
    curr_exp = sum(t["amount"] for t in monthly[curr] if t["type"] == "Expenses")
    prev_exp = sum(t["amount"] for t in monthly[prev] if t["type"] == "Expenses")
    change = ((curr_exp - prev_exp) / prev_exp * 100) if prev_exp else 0
    arrow = "⬆️" if change > 0 else "⬇️"
    
    print(f"\n📈 Month-over-Month:")
    print(f"  {prev} Expenses: ₹{prev_exp:,.0f}")
    print(f"  {curr} Expenses: ₹{curr_exp:,.0f}")
    print(f"  Change: {change:+.1f}% {arrow}")

# AI Insights using Ollama
report_data = f"""
Month: {latest_month}
Total Expenses: ₹{total_expense:,.0f}
Top categories: {dict(sorted(cat_totals.items(), key=lambda x: -x[1])[:5])}
Top merchants: {dict(sorted(merchant_totals.items(), key=lambda x: -x[1])[:5])}
Previous month expenses: ₹{prev_exp:,.0f}
Change: {change:+.1f}%
"""

print(f"\n🤖 AI Insights:")
try:
    response = requests.post("http://localhost:11434/api/generate", json={
        "model": "llama3.1:8b",
        "prompt": f"""You are a personal finance advisor for an Indian professional in Bengaluru.
Analyze this monthly spending data and give 3 short, specific, actionable insights in bullet points.
Be direct, use ₹ amounts, mention specific merchants/categories.

{report_data}

Respond with exactly 3 bullet points, each 1 line. No intro, no outro.""",
        "stream": False
    }, timeout=30)
    print(response.json()["response"])
except:
    print("  Ollama not available — skipping AI insights")